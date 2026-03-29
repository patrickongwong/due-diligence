#!/usr/bin/env python3
"""
Create professionally formatted Excel workbook from financial statement JSON.
Usage: python3 format_excel.py data.json output.xlsx

Reads financial_statements and operating_metrics from JSON, creates a 4-tab
Excel workbook with Wall Street IB formatting conventions.
Automatically detects bank vs. corporate IS structure.
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 format_excel.py data.json output.xlsx")
        sys.exit(1)

    json_path = sys.argv[1]
    excel_path = sys.argv[2]

    with open(json_path) as f:
        data = json.load(f)

    ticker = data.get("ticker", "UNKNOWN")
    company_name = data.get("company_name", ticker)
    fs = data.get("financial_statements", {})
    is_items = fs.get("income_statement", [])
    bs_items = fs.get("balance_sheet", [])
    cf_items = fs.get("cash_flow_statement", [])
    metrics = data.get("operating_metrics", {})

    # Determine periods
    def get_periods(items):
        p = set()
        if items:
            for item in items:
                p.update(k for k in item.keys() if k != "line_item" and k.isdigit())
        return sorted(p, reverse=True)

    periods_is = get_periods(is_items)
    periods_bs = get_periods(bs_items)
    periods_cf = get_periods(cf_items)
    periods_metrics = sorted(metrics.keys(), reverse=True) if metrics else []

    # ── Styles ────────────────────────────────────────────────────────────────
    FONT = "Arial"
    SZ = 10
    f_title = Font(name=FONT, size=12, bold=True)
    f_subtitle = Font(name=FONT, size=SZ, italic=True, color="666666")
    f_header = Font(name=FONT, size=SZ, bold=True)
    f_item = Font(name=FONT, size=SZ)
    f_bold = Font(name=FONT, size=SZ, bold=True)
    f_col_hdr = Font(name=FONT, size=SZ, bold=True)
    f_source = Font(name=FONT, size=8, italic=True, color="999999")
    fill_section = PatternFill("solid", fgColor="D9E1F2")
    a_left = Alignment(horizontal="left", vertical="center")
    a_indent1 = Alignment(horizontal="left", vertical="center", indent=2)
    a_center = Alignment(horizontal="center", vertical="center")
    a_right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin")
    med = Side(style="medium")
    dbl = Side(style="double")
    b_none = Border()
    b_top = Border(top=thin)
    b_bot_med = Border(bottom=med)
    b_top_dbl_bot = Border(top=thin, bottom=dbl)

    FMT_M = '#,##0;(#,##0);"-"'
    FMT_EPS = '$#,##0.00;($#,##0.00);"-"'
    FMT_SHARES = '#,##0.0'
    FMT_PCT = '0.0%;(0.0%);"-"'
    FMT_PCT2 = '0.00%;(0.00%);"-"'
    FMT_DLR = '$#,##0.00;($#,##0.00);"-"'
    FMT_X = '0.00"x";(0.00"x");"-"'

    # ── Helper functions ──────────────────────────────────────────────────────
    def lookup(items, label):
        if not items:
            return None
        for item in items:
            if item["line_item"] == label:
                return item
        return None

    def lookup_partial(items, partial, exclude=None):
        if not items:
            return None
        for item in items:
            lbl = item["line_item"]
            if partial.lower() in lbl.lower():
                if exclude and any(ex.lower() in lbl.lower() for ex in exclude):
                    continue
                return item
        return None

    def find(items, *labels):
        for label in labels:
            item = lookup(items, label)
            if item:
                return item
        return None

    def find_p(items, partial, **kwargs):
        return lookup_partial(items, partial, **kwargs)

    def get_val(item, period, to_millions=True, force_abs=False, negate=False):
        if item is None:
            return None
        v = item.get(period)
        if v is None or v == "":
            return None
        try:
            num = float(v)
            if to_millions:
                num = num / 1_000_000
            if force_abs:
                num = abs(num)
            if negate:
                num = -num
            return num
        except (ValueError, TypeError):
            return None

    def gm(year, key, scale=1.0):
        v = metrics.get(year, {}).get(key)
        return float(v) * scale if v is not None else None

    # ── Detect bank vs corporate ──────────────────────────────────────────────
    is_bank = any(
        any(kw in (item.get("line_item", "").lower()) for kw in
            ["interest and fees on finance receivables", "interest on deposits", "net interest income", "net financing revenue"])
        for item in (is_items or [])
    )
    print(f"Company type detected: {'Bank/Financial' if is_bank else 'Corporate'}")

    # ── Build structures ──────────────────────────────────────────────────────
    # Each entry: (row_type, label, source_item_or_getter, options)
    # options: {"abs": True, "negate": True, "fmt": "eps"|"shares", "no_millions": True}

    def build_is_structure():
        structure = []
        if is_bank:
            # Bank IS structure
            structure += [
                ("header", "Interest Income", None, {}),
                ("item", "Interest & fees on loans", find(is_items, "Interest and fees on finance receivables and loans"), {}),
                ("item", "Interest on loans held-for-sale", find(is_items, "Interest on loans held-for-sale"), {}),
                ("item", "Interest on investment securities", find_p(is_items, "Interest and dividends on investment securities"), {}),
                ("item", "Interest on cash and equivalents", find(is_items, "Interest on cash and cash equivalents"), {}),
                ("item", "Operating leases", find(is_items, "Operating leases"), {}),
                ("subtotal", "Total interest income", find(is_items, "Total financing revenue and other interest income") or find_p(is_items, "Total interest income"), {}),
                ("blank", "", None, {}),
                ("header", "Interest Expense", None, {}),
                ("item", "Interest on deposits", find(is_items, "Interest on deposits"), {}),
                ("item", "Interest on short-term borrowings", find(is_items, "Interest on short-term borrowings"), {}),
                ("item", "Interest on long-term debt", find(is_items, "Interest on long-term debt"), {}),
                ("subtotal", "Total interest expense", find(is_items, "Total interest expense"), {"abs": True}),
                ("blank", "", None, {}),
                ("total", "Net financing revenue", find(is_items, "Net financing revenue and other interest income") or find_p(is_items, "Net interest income"), {}),
                ("blank", "", None, {}),
            ]
            # Other revenue section if available
            other_rev = find(is_items, "Total other revenue") or find_p(is_items, "Total noninterest income")
            if other_rev:
                structure += [
                    ("header", "Other Revenue", None, {}),
                    ("item", "Insurance premiums and service revenue", find(is_items, "Insurance premiums and service revenue earned"), {}),
                    ("item", "Other income, net of losses", find(is_items, "Other income, net of losses"), {}),
                    ("item", "(Loss) gain on loans, net", find_p(is_items, "gain on mortgage and automotive loans"), {}),
                    ("item", "Other (loss) gain on investments", find_p(is_items, "Other (loss) gain on investments") or find_p(is_items, "Other gain (loss) on investments"), {}),
                    ("item", "Net depreciation on operating lease assets", find(is_items, "Net depreciation expense on operating lease assets"), {}),
                    ("subtotal", "Total other revenue", other_rev, {}),
                    ("blank", "", None, {}),
                ]
            structure += [
                ("total", "Total net revenue", find(is_items, "Total net revenue") or find_p(is_items, "Total revenue"), {}),
                ("blank", "", None, {}),
                ("item", "Provision for credit losses", find(is_items, "Provision for credit losses"), {"negate": True}),
                ("blank", "", None, {}),
            ]
        else:
            # Corporate IS structure
            structure += [
                ("header", "Revenue", None, {}),
                ("total", "Total revenue", find(is_items, "Revenue") or find(is_items, "Total revenue") or find_p(is_items, "revenue"), {}),
                ("item", "Cost of revenue", find_p(is_items, "cost of revenue") or find_p(is_items, "cost of goods sold"), {}),
                ("subtotal", "Gross profit", find_p(is_items, "gross profit"), {}),
                ("blank", "", None, {}),
            ]

        # Common expense section
        structure += [
            ("header", "Noninterest Expense" if is_bank else "Operating Expenses", None, {}),
            ("item", "Compensation and benefits", find(is_items, "Compensation and benefits expense") or find_p(is_items, "compensation"), {}),
        ]
        if is_bank:
            structure.append(("item", "Insurance losses", find(is_items, "Insurance losses and loss adjustment expenses"), {}))
            structure.append(("item", "Goodwill impairment", find(is_items, "Goodwill impairment"), {}))
        else:
            structure.append(("item", "Research and development", find_p(is_items, "research and development"), {}))
            structure.append(("item", "Selling, general and administrative", find_p(is_items, "selling, general"), {}))
        structure += [
            ("item", "Other operating expenses", find(is_items, "Other operating expenses") or find_p(is_items, "other operating"), {}),
            ("subtotal", "Total expense", find(is_items, "Total noninterest expense") or find_p(is_items, "total operating expense"), {"abs": True}),
            ("blank", "", None, {}),
        ]

        # Bottom of IS
        structure += [
            ("total", "Pre-tax income", find(is_items, "Income from continuing operations before income tax expense") or find_p(is_items, "income before income tax"), {}),
            ("item", "Income tax expense", find(is_items, "Total income tax expense from continuing operations") or find_p(is_items, "income tax expense"), {"abs": True}),
            ("blank", "", None, {}),
            ("total", "Net income from continuing operations", find(is_items, "Net income from continuing operations"), {}),
            ("item", "Discontinued operations", find(is_items, "Loss from discontinued operations, net of tax"), {}),
            ("grandtotal", "Net income", find(is_items, "Net income"), {}),
            ("blank", "", None, {}),
            ("item", "Preferred stock dividends", find_p(is_items, "Preferred stock dividends", exclude=["Series"]), {}),
            ("grandtotal", "Net income to common", find(is_items, "Net income attributable to common shareholders") or find_p(is_items, "net income attributable to common"), {}),
            ("blank", "", None, {}),
            ("header", "Per Share Data", None, {}),
            ("item", "Diluted EPS", find(is_items, "Net income (in dollars per share)") or find_p(is_items, "diluted earnings per common share"), {"fmt": "eps", "no_millions": True}),
            ("item", "Dividends per share", find(is_items, "Cash dividends declared per common share (in dollars per share)") or find(is_items, "Cash dividends declared per common share"), {"fmt": "eps", "no_millions": True}),
            ("item", "Diluted shares outstanding (mm)", find(is_items, "Diluted weighted-average common shares outstanding (in shares)") or find_p(is_items, "diluted weighted-average"), {"fmt": "shares"}),
        ]

        # Filter out entries where source is None (except headers/blanks)
        return [(t, l, s, o) for t, l, s, o in structure if t in ("header", "blank") or s is not None]

    # ── Write sheet function ──────────────────────────────────────────────────
    def write_financial_sheet(ws, title, subtitle, structure, periods):
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 48
        for i in range(len(periods)):
            ws.column_dimensions[get_column_letter(i + 3)].width = 15

        r = 1
        ws.cell(r, 2, company_name.upper()).font = f_title
        r += 1
        ws.cell(r, 2, subtitle).font = f_subtitle
        r += 2

        for i, p in enumerate(periods):
            c = ws.cell(r, i + 3, f"FY{p}")
            c.font = f_col_hdr
            c.alignment = a_center
            c.border = b_bot_med
        ws.cell(r, 2).border = b_bot_med
        r += 1

        for row_type, label, src, opts in structure:
            if row_type == "blank":
                r += 1
                continue
            lbl = ws.cell(r, 2, label)
            if row_type == "header":
                lbl.font = f_header
                lbl.alignment = a_left
                lbl.fill = fill_section
                for i in range(len(periods)):
                    ws.cell(r, i + 3).fill = fill_section
                r += 1
                continue

            is_bold_row = row_type in ("subtotal", "total", "grandtotal")
            lbl.font = f_bold if is_bold_row else f_item
            lbl.alignment = a_left if row_type in ("total", "grandtotal") else a_indent1

            if row_type == "subtotal":
                data_border = b_top
            elif row_type == "total":
                data_border = b_top
            elif row_type == "grandtotal":
                data_border = b_top_dbl_bot
            else:
                data_border = b_none

            fmt_key = opts.get("fmt", "millions")
            num_fmt = {"eps": FMT_EPS, "shares": FMT_SHARES}.get(fmt_key, FMT_M)
            no_millions = opts.get("no_millions", False)
            force_abs = opts.get("abs", False)
            negate = opts.get("negate", False)

            for i, period in enumerate(periods):
                cell = ws.cell(r, i + 3)
                if src is None:
                    cell.value = None
                elif fmt_key == "eps":
                    v = src.get(period)
                    try:
                        cell.value = float(v) if v is not None and v != "" else None
                    except:
                        cell.value = None
                elif fmt_key == "shares":
                    v = src.get(period)
                    try:
                        cell.value = float(v) / 1_000_000 if v is not None and v != "" else None
                    except:
                        cell.value = None
                else:
                    cell.value = get_val(src, period, to_millions=not no_millions, force_abs=force_abs, negate=negate)
                cell.number_format = num_fmt
                cell.alignment = a_right
                cell.font = f_bold if is_bold_row else f_item
                cell.border = data_border
            r += 1

        r += 2
        ws.cell(r, 2, f"Source: SEC EDGAR 10-K filings via edgartools").font = f_source

    # ── Write metrics sheet ───────────────────────────────────────────────────
    def write_metrics_sheet(ws, periods):
        if not metrics:
            return

        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 42
        for i in range(len(periods)):
            ws.column_dimensions[get_column_letter(i + 3)].width = 14

        r = 1
        ws.cell(r, 2, company_name.upper()).font = f_title
        r += 1
        ws.cell(r, 2, "Operating Metrics").font = f_subtitle
        r += 2

        for i, p in enumerate(periods):
            c = ws.cell(r, i + 3, f"FY{p}")
            c.font = f_col_hdr
            c.alignment = a_center
            c.border = b_bot_med
        ws.cell(r, 2).border = b_bot_med
        r += 1

        # Build metrics structure dynamically based on what's available
        # Check what metrics exist across all years
        all_metric_keys = set()
        for ym in metrics.values():
            all_metric_keys.update(ym.keys())

        def has_metric(key):
            return key in all_metric_keys

        entries = []

        # Profitability
        entries.append(("header", "Profitability"))
        if has_metric("roe"):
            entries.append(("item", "Return on avg equity (ROE)", "roe", FMT_PCT, 0.01))
        if has_metric("roa"):
            entries.append(("item", "Return on avg assets (ROA)", "roa", FMT_PCT2, 0.01))
        if has_metric("rotce"):
            entries.append(("item", "Return on avg TCE (ROTCE)", "rotce", FMT_PCT, 0.01))
        if has_metric("nim_pct"):
            entries.append(("item", "Net interest margin (NIM)", "nim_pct", FMT_PCT2, 0.01))
        if has_metric("efficiency_ratio"):
            entries.append(("item", "Efficiency ratio", "efficiency_ratio", FMT_PCT, 0.01))
        if has_metric("effective_tax_rate"):
            entries.append(("item", "Effective tax rate", "effective_tax_rate", FMT_PCT, 0.01))
        entries.append(("blank",))

        # Earnings Power
        if has_metric("ppnr_mm"):
            entries.append(("header", "Earnings Power ($mm)"))
            entries.append(("item", "Net interest income", "net_interest_income_mm", FMT_M, 1))
            entries.append(("item", "Noninterest income", "noninterest_income_mm", FMT_M, 1))
            entries.append(("subtotal", "Total net revenue", "total_net_revenue_mm", FMT_M, 1))
            entries.append(("item", "Noninterest expense", "noninterest_expense_mm", FMT_M, 1))
            entries.append(("subtotal", "Pre-provision net revenue (PPNR)", "ppnr_mm", FMT_M, 1))
            if has_metric("ppnr_to_assets"):
                entries.append(("item", "PPNR / avg assets", "ppnr_to_assets", FMT_PCT2, 0.01))
            entries.append(("blank",))

        # Yields & Spreads
        if has_metric("yield_on_earning_assets"):
            entries.append(("header", "Yields & Spreads"))
            entries.append(("item", "Yield on earning assets", "yield_on_earning_assets", FMT_PCT2, 0.01))
            if has_metric("yield_on_loans"):
                entries.append(("item", "Yield on loans", "yield_on_loans", FMT_PCT2, 0.01))
            if has_metric("cost_of_deposits"):
                entries.append(("item", "Cost of deposits", "cost_of_deposits", FMT_PCT2, 0.01))
            if has_metric("cost_of_funds"):
                entries.append(("item", "Cost of funds", "cost_of_funds", FMT_PCT2, 0.01))
            if has_metric("net_interest_spread"):
                entries.append(("item", "Net interest spread", "net_interest_spread", FMT_PCT2, 0.01))
            entries.append(("blank",))

        # Credit Quality
        if has_metric("provision_mm") or has_metric("net_chargeoffs_mm"):
            entries.append(("header", "Credit Quality"))
            if has_metric("gross_chargeoffs_mm"):
                entries.append(("item", "Gross charge-offs ($mm)", "gross_chargeoffs_mm", FMT_M, 1))
            if has_metric("recoveries_mm"):
                entries.append(("item", "Recoveries ($mm)", "recoveries_mm", FMT_M, 1))
            if has_metric("net_chargeoffs_mm"):
                entries.append(("subtotal", "Net charge-offs ($mm)", "net_chargeoffs_mm", FMT_M, 1))
            if has_metric("nco_ratio"):
                entries.append(("item", "NCO ratio", "nco_ratio", FMT_PCT2, 0.01))
            entries.append(("item", "Provision ($mm)", "provision_mm", FMT_M, 1))
            if has_metric("provision_to_loans"):
                entries.append(("item", "Provision / avg loans", "provision_to_loans", FMT_PCT2, 0.01))
            if has_metric("allowance_to_loans"):
                entries.append(("item", "Allowance / total loans", "allowance_to_loans", FMT_PCT2, 0.01))
            entries.append(("blank",))

        # Regulatory Capital
        if has_metric("cet1_ratio"):
            entries.append(("header", "Regulatory Capital"))
            entries.append(("item", "CET1 capital ($mm)", "cet1_capital_mm", FMT_M, 1))
            entries.append(("item", "CET1 ratio", "cet1_ratio", FMT_PCT2, 0.01))
            if has_metric("tier1_ratio"):
                entries.append(("item", "Tier 1 capital ($mm)", "tier1_capital_mm", FMT_M, 1))
                entries.append(("item", "Tier 1 ratio", "tier1_ratio", FMT_PCT2, 0.01))
            if has_metric("total_capital_ratio"):
                entries.append(("item", "Total capital ratio", "total_capital_ratio", FMT_PCT2, 0.01))
            entries.append(("blank",))

        # Balance Sheet
        if has_metric("total_assets_mm"):
            entries.append(("header", "Balance Sheet ($mm)"))
            entries.append(("item", "Total assets", "total_assets_mm", FMT_M, 1))
            if has_metric("total_loans_net_mm"):
                entries.append(("item", "Total loans, net", "total_loans_net_mm", FMT_M, 1))
            if has_metric("total_deposits_mm"):
                entries.append(("item", "Total deposits", "total_deposits_mm", FMT_M, 1))
            entries.append(("item", "Total equity", "total_equity_mm", FMT_M, 1))
            entries.append(("blank",))

        # Funding & Leverage
        if has_metric("loan_to_deposit") or has_metric("leverage_ratio"):
            entries.append(("header", "Funding & Leverage"))
            if has_metric("loan_to_deposit"):
                entries.append(("item", "Loan-to-deposit ratio", "loan_to_deposit", FMT_PCT, 0.01))
            if has_metric("deposits_to_liabilities"):
                entries.append(("item", "Deposits / total liabilities", "deposits_to_liabilities", FMT_PCT, 0.01))
            if has_metric("leverage_ratio"):
                entries.append(("item", "Equity / total assets", "leverage_ratio", FMT_PCT, 0.01))
            entries.append(("blank",))

        # Insurance & Leasing
        if has_metric("insurance_loss_ratio"):
            entries.append(("header", "Insurance & Leasing"))
            entries.append(("item", "Insurance loss ratio", "insurance_loss_ratio", FMT_PCT, 0.01))
            if has_metric("operating_lease_net_revenue"):
                entries.append(("item", "Op. lease net revenue ($mm)", "operating_lease_net_revenue", FMT_M, 0.000001))
            entries.append(("blank",))

        # Per Share
        entries.append(("header", "Per Share Data"))
        if has_metric("eps_diluted"):
            entries.append(("item", "Diluted EPS", "eps_diluted", FMT_DLR, 1))
        if has_metric("dps"):
            entries.append(("item", "Dividends per share", "dps", FMT_DLR, 1))
        if has_metric("bvps"):
            entries.append(("item", "Book value per share", "bvps", FMT_DLR, 1))
        if has_metric("tbvps"):
            entries.append(("item", "Tangible book value per share", "tbvps", FMT_DLR, 1))
        entries.append(("blank",))

        # Capital Return
        if has_metric("buybacks_mm") or has_metric("dividend_payout_ratio"):
            entries.append(("header", "Capital Return"))
            if has_metric("buybacks_mm"):
                entries.append(("item", "Share repurchases ($mm)", "buybacks_mm", FMT_M, 1))
            if has_metric("dividend_payout_ratio"):
                entries.append(("item", "Dividend payout ratio", "dividend_payout_ratio", FMT_PCT, 0.01))
            if has_metric("total_payout_ratio"):
                entries.append(("item", "Total payout ratio", "total_payout_ratio", FMT_PCT, 0.01))
            entries.append(("blank",))

        # Share Data
        if has_metric("diluted_shares_mm"):
            entries.append(("header", "Share Data (millions)"))
            entries.append(("item", "Diluted shares", "diluted_shares_mm", FMT_SHARES, 1))
            if has_metric("shares_outstanding_mm"):
                entries.append(("item", "Basic shares", "shares_outstanding_mm", FMT_SHARES, 1))

        # Write entries
        for entry in entries:
            if entry[0] == "blank":
                r += 1
                continue

            row_type = entry[0]
            label = entry[1]

            if row_type == "header":
                lbl_cell = ws.cell(r, 2, label)
                lbl_cell.font = f_header
                lbl_cell.alignment = a_left
                lbl_cell.fill = fill_section
                for i in range(len(periods)):
                    ws.cell(r, i + 3).fill = fill_section
                r += 1
                continue

            key, fmt, scale = entry[2], entry[3], entry[4]
            is_bold_row = row_type in ("subtotal",)
            lbl_cell = ws.cell(r, 2, label)
            lbl_cell.font = f_bold if is_bold_row else f_item
            lbl_cell.alignment = a_left if is_bold_row else a_indent1
            data_border = b_top if is_bold_row else b_none

            for i, year in enumerate(periods):
                cell = ws.cell(r, i + 3)
                cell.value = gm(year, key, scale)
                cell.number_format = fmt
                cell.alignment = a_right
                cell.font = f_bold if is_bold_row else f_item
                cell.border = data_border
            r += 1

        r += 2
        ws.cell(r, 2, "Source: SEC EDGAR 10-K XBRL data via edgartools").font = f_source

    # ── Build and write all sheets ────────────────────────────────────────────
    wb = Workbook()

    # Income Statement
    ws_is = wb.active
    ws_is.title = "Income Statement"
    is_structure = build_is_structure()
    write_financial_sheet(ws_is, company_name, "Consolidated Statement of Income  ($ in millions, except per share)", is_structure, periods_is)

    # Balance Sheet — use all BS items directly (dynamic)
    # For BS we output all non-abstract, non-empty line items grouped logically
    ws_bs = wb.create_sheet("Balance Sheet")
    bs_structure = []

    # Assets
    bs_structure.append(("header", "Assets", None, {}))
    for label_search in [
        "Total cash and cash equivalents", "Cash and cash equivalents",
    ]:
        item = find(bs_items, label_search)
        if item:
            bs_structure.append(("item", "Cash and cash equivalents", item, {}))
            break

    for label_search in ["Available-for-sale securities"]:
        item = find_p(bs_items, label_search)
        if item:
            bs_structure.append(("item", "Available-for-sale securities", item, {}))
            break

    for label_search in ["Held-to-maturity securities"]:
        item = find_p(bs_items, label_search)
        if item:
            bs_structure.append(("item", "Held-to-maturity securities", item, {}))
            break

    for item_def in [
        ("Equity securities", find(bs_items, "Equity securities")),
        ("Loans held-for-sale, net", find(bs_items, "Loans held-for-sale, net")),
        ("Gross loans", find(bs_items, "Finance receivables and loans, net of unearned income") or find_p(bs_items, "gross loans")),
        ("Allowance for loan losses", find(bs_items, "Allowance for loan losses") or find_p(bs_items, "allowance for credit losses")),
        ("Total loans, net", find(bs_items, "Total finance receivables and loans, net") or find_p(bs_items, "total loans, net")),
        ("Investment in operating leases", find(bs_items, "Investment in operating leases, net")),
        ("Property and equipment", find_p(bs_items, "property and equipment") or find_p(bs_items, "property, plant")),
        ("Goodwill", find_p(bs_items, "goodwill")),
        ("Insurance assets", find(bs_items, "Premiums receivable and other insurance assets")),
        ("Other assets", find(bs_items, "Other assets")),
    ]:
        if item_def[1]:
            row_type = "subtotal" if "Total" in item_def[0] else "item"
            bs_structure.append((row_type, item_def[0], item_def[1], {}))

    bs_structure.append(("total", "Total assets", find(bs_items, "Total assets"), {}))
    bs_structure.append(("blank", "", None, {}))

    # Liabilities
    bs_structure.append(("header", "Liabilities", None, {}))
    for item_def in [
        ("Total deposit liabilities", find(bs_items, "Total deposit liabilities") or find_p(bs_items, "total deposits")),
        ("Short-term borrowings", find(bs_items, "Short-term borrowings")),
        ("Long-term debt", find(bs_items, "Long-term debt")),
        ("Interest payable", find(bs_items, "Interest payable")),
        ("Accounts payable", find_p(bs_items, "accounts payable")),
        ("Unearned premiums", find(bs_items, "Unearned insurance premiums and service revenue")),
        ("Accrued expenses and other", find(bs_items, "Accrued expenses and other liabilities") or find_p(bs_items, "accrued expenses")),
    ]:
        if item_def[1]:
            bs_structure.append(("item", item_def[0], item_def[1], {}))

    bs_structure.append(("total", "Total liabilities", find(bs_items, "Total liabilities"), {}))
    bs_structure.append(("blank", "", None, {}))

    # Equity
    bs_structure.append(("header", "Shareholders' Equity", None, {}))
    for item_def in [
        ("Preferred stock", find(bs_items, "Preferred stock")),
        ("Common stock and paid-in capital", find_p(bs_items, "Common stock and paid-in capital")),
        ("Retained earnings", find(bs_items, "Retained earnings")),
        ("AOCI", find(bs_items, "Accumulated other comprehensive loss") or find_p(bs_items, "accumulated other comprehensive")),
        ("Treasury stock", find_p(bs_items, "Treasury stock, at cost")),
    ]:
        if item_def[1]:
            bs_structure.append(("item", item_def[0], item_def[1], {}))

    bs_structure.append(("total", "Total equity", find(bs_items, "Total equity") or find_p(bs_items, "total stockholders"), {}))
    bs_structure.append(("blank", "", None, {}))
    bs_structure.append(("grandtotal", "Total liabilities and equity", find(bs_items, "Total liabilities and equity"), {}))

    # Filter out None sources
    bs_structure = [(t, l, s, o) for t, l, s, o in bs_structure if t in ("header", "blank") or s is not None]
    write_financial_sheet(ws_bs, company_name, "Consolidated Balance Sheet  ($ in millions)", bs_structure, periods_bs)

    # Cash Flow Statement
    ws_cf = wb.create_sheet("Cash Flow Statement")
    cf_structure = [
        ("header", "Operating Activities", None, {}),
        ("item", "Net income", find(cf_items, "Net income"), {}),
        ("item", "Depreciation and amortization", find(cf_items, "Depreciation and amortization"), {}),
        ("item", "Provision for credit losses", find(cf_items, "Provision for credit losses"), {}),
        ("item", "Deferred income taxes", find(cf_items, "Deferred income taxes"), {}),
        ("item", "Stock-based compensation", find_p(cf_items, "stock-based compensation") or find_p(cf_items, "share-based compensation"), {}),
        ("item", "Goodwill impairment", find(cf_items, "Goodwill impairment"), {}),
        ("item", "Other adjustments", find(cf_items, "Other, net"), {}),
        ("subtotal", "Net cash from operations", find(cf_items, "Net cash provided by operating activities") or find_p(cf_items, "net cash provided by operating") or find_p(cf_items, "cash from operating"), {}),
        ("blank", "", None, {}),
        ("header", "Investing Activities", None, {}),
        ("item", "Purchases of securities", find(cf_items, "Purchases of available-for-sale securities") or find_p(cf_items, "purchases of investments"), {}),
        ("item", "Proceeds from securities", find(cf_items, "Proceeds from sales of available-for-sale securities"), {}),
        ("item", "Purchases of loans/receivables", find(cf_items, "Purchases of finance receivables and loans held-for-investment"), {}),
        ("item", "Capital expenditures", find_p(cf_items, "purchases of property") or find_p(cf_items, "capital expenditure"), {}),
        ("item", "Purchases of operating lease assets", find(cf_items, "Purchases of operating lease assets"), {}),
        ("item", "Disposals of operating lease assets", find(cf_items, "Disposals of operating lease assets"), {}),
        ("item", "Acquisitions/divestitures, net", find(cf_items, "Proceeds from sale of operation or business unit, net") or find_p(cf_items, "acquisitions"), {}),
        ("subtotal", "Net cash from investing", find(cf_items, "Net cash (used in) provided by investing activities") or find_p(cf_items, "net cash provided by (used in) investing") or find_p(cf_items, "net cash used in investing"), {}),
        ("blank", "", None, {}),
        ("header", "Financing Activities", None, {}),
        ("item", "Net change in deposits", find(cf_items, "Net (decrease) increase in deposits") or find_p(cf_items, "net increase in deposits"), {}),
        ("item", "Net change in borrowings", find(cf_items, "Net change in short-term borrowings"), {}),
        ("item", "Proceeds from debt issuance", find(cf_items, "Proceeds from issuance of long-term debt"), {}),
        ("item", "Repayments of debt", find(cf_items, "Repayments of long-term debt"), {}),
        ("item", "Common dividends paid", find(cf_items, "Common stock dividends paid") or find_p(cf_items, "dividends paid"), {}),
        ("item", "Preferred dividends paid", find(cf_items, "Preferred stock dividends paid"), {}),
        ("item", "Share repurchases", find(cf_items, "Repurchases of common stock") or find_p(cf_items, "repurchase of common stock"), {}),
        ("subtotal", "Net cash from financing", find(cf_items, "Net cash provided by (used in) financing activities") or find_p(cf_items, "net cash (used in) provided by financing") or find_p(cf_items, "net cash used in financing"), {}),
        ("blank", "", None, {}),
        ("grandtotal", "Free cash flow", find(cf_items, "Free cash flow"), {}),
        ("blank", "", None, {}),
        ("grandtotal", "Net change in cash", find(cf_items, "Net increase in cash and cash equivalents and restricted cash") or find_p(cf_items, "net increase (decrease) in cash"), {}),
        ("blank", "", None, {}),
        ("total", "Cash, end of period", find_p(cf_items, "Total cash and cash equivalents and restricted cash") or find_p(cf_items, "cash and cash equivalents, end"), {}),
    ]
    cf_structure = [(t, l, s, o) for t, l, s, o in cf_structure if t in ("header", "blank") or s is not None]
    write_financial_sheet(ws_cf, company_name, "Consolidated Statement of Cash Flows  ($ in millions)", cf_structure, periods_cf)

    # Operating Metrics
    if metrics:
        ws_om = wb.create_sheet("Operating Metrics")
        write_metrics_sheet(ws_om, periods_metrics)

    # Print settings
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(excel_path)
    print(f"Saved: {excel_path}")
    print(f"Sheets: {wb.sheetnames}")
    print("FORMAT_COMPLETE")


if __name__ == "__main__":
    main()
